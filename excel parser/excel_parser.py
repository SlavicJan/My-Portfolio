#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

import sys, os, re
from pathlib import Path
from typing import Optional, Tuple, Dict, Any, List
import logging
from logging import handlers
import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

import tkinter as tk
from tkinter import filedialog, messagebox

LOG_FILENAME = "excel_parser.log"
logger = logging.getLogger("excel_parser")
logger.setLevel(logging.INFO)
ch = logging.StreamHandler()
ch.setLevel(logging.INFO)
ch.setFormatter(logging.Formatter("%(asctime)s %(levelname)s: %(message)s", "%Y-%m-%d %H:%M:%S"))
logger.addHandler(ch)
fh = handlers.RotatingFileHandler(LOG_FILENAME, maxBytes=5*1024*1024, backupCount=3, encoding="utf-8")
fh.setLevel(logging.INFO)
fh.setFormatter(logging.Formatter("%(asctime)s %(levelname)s: %(message)s", "%Y-%m-%d %H:%M:%S"))
logger.addHandler(fh)

KEYS = {
    "account_label": ["выписка по счёту", "выписка по счету", "выписка по сч", "счёт", "счет"],
    "period": ["за период", "период"],
    "owner": ["владелец счета", "владелец счёта", "владелец", "владельца"],
    "inn": ["инн владельца", "инн", "инн владельца:"],
    "bik": ["бик"],
    "incoming": ["остаток входящий", "входящий остаток", "остаток на начало", "входящий остаток:"],
    "outgoing": ["остаток исходящий", "исходящий остаток", "остаток на конец", "исходящий остаток:"],
    "turnover_debit": ["обороты дебет", "оборот дебет", "списание", "сумма списание"],
    "turnover_credit": ["обороты кредит", "оборот кредит", "поступление", "сумма поступление"],
    "turnover_header": ["обороты за период", "обороты за", "оборот за период", "обороты"]
}

def norm(s: Any) -> str:
    if s is None:
        return ""
    return str(s).strip()

def cell_text(cell: Cell) -> str:
    return norm(cell.value).lower()

def normalize_account(s: Optional[str]) -> Optional[str]:
    if s is None:
        return None
    t = str(s).replace("\xa0", " ")
    t = re.sub(r"\s+", " ", t)
    return t.strip()

def parse_amount(s: Optional[str]) -> Optional[float]:
    if s is None:
        return None
    t = str(s).strip().replace("\xa0", "").replace(" ", "")
    t = t.replace(",", ".")
    neg = False
    if t.startswith("(") and t.endswith(")"):
        neg = True
        t = t[1:-1]
    m = re.search(r"-?\d+(\.\d+)?", t)
    if not m:
        try:
            val = float(t)
            return -val if neg else val
        except Exception:
            return None
    try:
        val = float(m.group(0))
        return -val if neg else val
    except Exception:
        return None

def find_cell_by_phrase(sheet, phrases: List[str]):
    for row in sheet.iter_rows(values_only=False):
        for cell in row:
            txt = cell_text(cell)
            if not txt:
                continue
            for p in phrases:
                if p in txt:
                    return cell
    return None

def find_all_cells_by_phrase(sheet, phrases: List[str]) -> List:
    res=[]
    for row in sheet.iter_rows(values_only=False):
        for cell in row:
            txt = cell_text(cell)
            if not txt: continue
            for p in phrases:
                if p in txt:
                    res.append(cell); break
    return res

def right_neighbor_value(sheet, cell, max_offset=6):
    if cell is None: return None
    for offset in range(1, max_offset+1):
        c = sheet.cell(row=cell.row, column=cell.column+offset)
        if c and c.value not in (None, ""):
            return norm(c.value)
    return None

def left_neighbor_value(sheet, cell, max_offset=4):
    if cell is None: return None
    for offset in range(1, max_offset+1):
        col = cell.column - offset
        if col < 1: break
        c = sheet.cell(row=cell.row, column=col)
        if c and c.value not in (None, ""):
            return norm(c.value)
    return None

def search_in_columns(sheet, cols, phrases):
    res=[]
    for col in cols:
        for row in sheet.iter_rows(min_col=col, max_col=col, values_only=False):
            c = row[0]; txt = cell_text(c)
            if not txt: continue
            for p in phrases:
                if p in txt:
                    res.append(c); break
    return res

def find_transactions_table(sheet):
    max_row=sheet.max_row
    for r in range(1, max_row+1):
        a = sheet.cell(row=r,column=1).value
        c = sheet.cell(row=r,column=3).value
        d = sheet.cell(row=r,column=4).value
        if a is None: continue
        if (isinstance(c,(int,float)) or isinstance(d,(int,float))) and (isinstance(a,str) or hasattr(a,"year") or isinstance(a,(int,float))):
            start=r; end=r
            for rr in range(r+1, max_row+1):
                if sheet.cell(row=rr,column=1).value in (None,""): break
                end=rr
            return (start,end)
    return None

def extract_by_first_template(sheet):
    res={}
    c=find_cell_by_phrase(sheet, KEYS["account_label"])
    if c:
        res["метка_счёта"]=norm(c.value)
        acc = right_neighbor_value(sheet,c) or left_neighbor_value(sheet,c)
        res["номер_счёта"]=normalize_account(acc) if acc else None
    c=find_cell_by_phrase(sheet, KEYS["period"])
    if c: res["период"]= right_neighbor_value(sheet,c) or left_neighbor_value(sheet,c)
    c=find_cell_by_phrase(sheet, KEYS["owner"])
    if c: res["владелец"]= right_neighbor_value(sheet,c) or left_neighbor_value(sheet,c)
    c=find_cell_by_phrase(sheet, KEYS["inn"])
    if c: res["инн"]= right_neighbor_value(sheet,c) or left_neighbor_value(sheet,c)
    c=find_cell_by_phrase(sheet, KEYS["bik"])
    if c: res["бик"]= right_neighbor_value(sheet,c) or left_neighbor_value(sheet,c)

    found_inc = search_in_columns(sheet, [1,2,3,4], KEYS["incoming"])
    inc = (right_neighbor_value(sheet, found_inc[0]) or left_neighbor_value(sheet, found_inc[0])) if found_inc else None
    found_out = search_in_columns(sheet, [1,2,3,4], KEYS["outgoing"])
    out = (right_neighbor_value(sheet, found_out[-1]) or left_neighbor_value(sheet, found_out[-1])) if found_out else None

    res["остаток_входящий_стр"]=inc
    res["остаток_исходящий_стр"]=out
    res["остаток_входящий"]=parse_amount(inc)
    res["остаток_исходящий"]=parse_amount(out)

    td_cells = find_all_cells_by_phrase(sheet, KEYS["turnover_debit"])
    tc_cells = find_all_cells_by_phrase(sheet, KEYS["turnover_credit"])
    td = (right_neighbor_value(sheet, td_cells[0]) or left_neighbor_value(sheet, td_cells[0])) if td_cells else None
    tc = (right_neighbor_value(sheet, tc_cells[0]) or left_neighbor_value(sheet, tc_cells[0])) if tc_cells else None

    res["об_дт_стр"]=td
    res["об_кт_стр"]=tc
    res["об_дт"]=parse_amount(td)
    res["об_кт"]=parse_amount(tc)
    return res

def extract_by_second_template(sheet):
    res={}
    header_vals=[norm(sheet.cell(row=1,column=col).value) for col in range(1,5)]
    header_line=" ".join([v for v in header_vals if v])
    res["заголовок_стр"]= header_line if header_line else None
    if header_line and any(p in header_line.lower() for p in KEYS["account_label"]):
        acc_match = re.search(r"\d{6,}", header_line.replace("–","-"))
        res["номер_счёта"]= normalize_account(acc_match.group(0)) if acc_match else None
        period_match = re.search(r"с\s*([^,;]+?)\s*по\s*([^\s,;]+)", header_line, flags=re.IGNORECASE)
        if period_match: res["период"]= period_match.group(1).strip()+" — "+period_match.group(2).strip()

    incoming_cell=None
    try:
        c5 = sheet.cell(row=5,column=1)
        if c5.value and any(k in cell_text(c5) for k in KEYS["incoming"]): incoming_cell=c5
    except Exception:
        incoming_cell=None
    if not incoming_cell:
        found = search_in_columns(sheet, [1,2,3,4], KEYS["incoming"])
        incoming_cell = found[0] if found else None
    res["остаток_входящий_стр"]= right_neighbor_value(sheet,incoming_cell) if incoming_cell else None
    res["остаток_входящий"]= parse_amount(res["остаток_входящий_стр"])

    found_out = search_in_columns(sheet, [1,2,3,4], KEYS["outgoing"])
    if found_out:
        last_out = found_out[-1]
        res["остаток_исходящий_стр"]= right_neighbor_value(sheet,last_out) or left_neighbor_value(sheet,last_out)
    else:
        res["остаток_исходящий_стр"]= None
    res["остаток_исходящий"]= parse_amount(res["остаток_исходящий_стр"])

    found_turn_header = find_all_cells_by_phrase(sheet, KEYS["turnover_header"])
    turnover_debit=None; turnover_credit=None
    if found_turn_header:
        h=found_turn_header[0]
        c3=sheet.cell(row=h.row,column=3); c4=sheet.cell(row=h.row,column=4)
        if c3 and c3.value not in (None,""): turnover_debit=norm(c3.value)
        if c4 and c4.value not in (None,""): turnover_credit=norm(c4.value)
        if turnover_debit is None:
            for rr in range(h.row, sheet.max_row+1):
                v=sheet.cell(row=rr,column=3).value
                if v not in (None,""): turnover_debit=norm(v); break
        if turnover_credit is None:
            for rr in range(h.row, sheet.max_row+1):
                v=sheet.cell(row=rr,column=4).value
                if v not in (None,""): turnover_credit=norm(v); break
    else:
        found_deb = search_in_columns(sheet, [3], KEYS["turnover_debit"])
        found_cred = search_in_columns(sheet, [4], KEYS["turnover_credit"])
        if found_deb: turnover_debit = right_neighbor_value(sheet, found_deb[0]) or left_neighbor_value(sheet, found_deb[0])
        if found_cred: turnover_credit = right_neighbor_value(sheet, found_cred[0]) or left_neighbor_value(sheet, found_cred[0])

    res["об_дт_стр"]= turnover_debit
    res["об_кт_стр"]= turnover_credit
    res["об_дт"]= parse_amount(turnover_debit)
    res["об_кт"]= parse_amount(turnover_credit)

    if res["об_кт"] is not None and res["об_дт"] is not None:
        res["выручка_по_оборотам"] = res["об_кт"] - res["об_дт"]
    else:
        res["выручка_по_оборотам"] = None
    if res["остаток_исходящий"] is not None and res["остаток_входящий"] is not None:
        res["изменение_остатка"] = res["остаток_исходящий"] - res["остаток_входящий"]
    else:
        res["изменение_остатка"] = None
    if res["выручка_по_оборотам"] is not None and res["изменение_остатка"] is not None:
        res["расхождение_сверки"] = res["выручка_по_оборотам"] - res["изменение_остатка"]
    else:
        res["расхождение_сверки"] = None
    return res

def extract_from_file(path: Path) -> Dict[str, Any]:
    logger.info("Обрабатываю: %s", path.name)
    base = {
        "Файл": str(path),
        "Лист": None,
        "Метка счета": None,
        "Номер счета": None,
        "Период": None,
        "Владелец": None,
        "ИНН": None,
        "БИК": None,
        "Остаток входящий (raw)": None,
        "Остаток исходящий (raw)": None,
        "Остаток входящий": None,
        "Остаток исходящий": None,
        "Обороты Дт (raw)": None,
        "Обороты Кт (raw)": None,
        "Обороты Дт": None,
        "Обороты Кт": None,
        "Выручка по оборотам": None,
        "Изменение остатка": None,
        "Расхождение сверки": None,
        "Транзакции найдены": False,
        "Лист транзакций": None
    }
    try:
        wb = load_workbook(filename=str(path), data_only=True)
    except Exception as e:
        logger.warning("Не открылся %s: %s", path.name, e)
        base["error"]=str(e); return base

    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        base["Лист"]=sheetname

        first = extract_by_first_template(sheet)
        if first.get("метка_счёта"): base["Метка счета"]=first.get("метка_счёта")
        if first.get("номер_счёта"): base["Номер счета"]=first.get("номер_счёта")
        if first.get("период"): base["Период"]=first.get("период")
        if first.get("владелец"): base["Владелец"]=first.get("владелец")
        if first.get("инн"): base["ИНН"]=first.get("инн")
        if first.get("бик"): base["БИК"]=first.get("бик")
        if first.get("остаток_входящий_стр") is not None: base["Остаток входящий (raw)"]=first.get("остаток_входящий_стр")
        if first.get("остаток_исходящий_стр") is not None: base["Остаток исходящий (raw)"]=first.get("остаток_исходящий_стр")
        if first.get("остаток_входящий") is not None: base["Остаток входящий"]=first.get("остаток_входящий")
        if first.get("остаток_исходящий") is not None: base["Остаток исходящий"]=first.get("остаток_исходящий")
        if first.get("об_дт_стр") is not None: base["Обороты Дт (raw)"]=first.get("об_дт_стр")
        if first.get("об_кт_стр") is not None: base["Обороты Кт (raw)"]=first.get("об_кт_стр")
        if first.get("об_дт") is not None: base["Обороты Дт"]=first.get("об_дт")
        if first.get("об_кт") is not None: base["Обороты Кт"]=first.get("об_кт")

        key_found = any([base.get("Номер счета"), base.get("Остаток входящий"), base.get("Остаток исходящий"),
                         base.get("Обороты Дт"), base.get("Обороты Кт")])
        if not key_found:
            second = extract_by_second_template(sheet)
            if second.get("номер_счёта"): base["Номер счета"]=second.get("номер_счёта")
            if second.get("период"): base["Период"]=second.get("период")
            if second.get("остаток_входящий_стр") is not None: base["Остаток входящий (raw)"]=second.get("остаток_входящий_стр")
            if second.get("остаток_исходящий_стр") is not None: base["Остаток исходящий (raw)"]=second.get("остаток_исходящий_стр")
            if second.get("остаток_входящий") is not None: base["Остаток входящий"]=second.get("остаток_входящий")
            if second.get("остаток_исходящий") is not None: base["Остаток исходящий"]=second.get("остаток_исходящий")
            if second.get("об_дт_стр") is not None: base["Обороты Дт (raw)"]=second.get("об_дт_стр")
            if second.get("об_кт_стр") is not None: base["Обороты Кт (raw)"]=second.get("об_кт_стр")
            if second.get("об_дт") is not None: base["Обороты Дт"]=second.get("об_дт")
            if second.get("об_кт") is not None: base["Обороты Кт"]=second.get("об_кт")
            if second.get("выручка_по_оборотам") is not None: base["Выручка по оборотам"]=second.get("выручка_по_оборотам")
            if second.get("изменение_остатка") is not None: base["Изменение остатка"]=second.get("изменение_остатка")
            if second.get("расхождение_сверки") is not None: base["Расхождение сверки"]=second.get("расхождение_сверки")

        tt = find_transactions_table(sheet)
        if tt:
            base["Транзакции найдены"]=True
            start,end = tt
            rows=[]
            for r in range(start, end+1):
                date = sheet.cell(row=r,column=1).value
                desc = sheet.cell(row=r,column=5).value
                debit = sheet.cell(row=r,column=3).value
                credit = sheet.cell(row=r,column=4).value
                if (debit in (None,"") and credit in (None,"")):
                    alt_debit = sheet.cell(row=r,column=4).value
                    alt_credit = sheet.cell(row=r,column=5).value
                    debit = debit or alt_debit
                    credit = credit or alt_credit
                rows.append({
                    "Дата": norm(date),
                    "Описание": norm(desc),
                    "Списание": parse_amount(norm(debit)) if debit not in (None,"") else None,
                    "Поступление": parse_amount(norm(credit)) if credit not in (None,"") else None
                })
            base["Транзакции_DF"]= pd.DataFrame(rows)
            break

    try:
        if base.get("Выручка по оборотам") is None:
            if base.get("Обороты Кт") is not None and base.get("Обороты Дт") is not None:
                base["Выручка по оборотам"] = base["Обороты Кт"] - base["Обороты Дт"]
        if base.get("Изменение остатка") is None:
            if base.get("Остаток исходящий") is not None and base.get("Остаток входящий") is not None:
                base["Изменение остатка"] = base["Остаток исходящий"] - base["Остаток входящий"]
        if base.get("Расхождение сверки") is None:
            if base.get("Выручка по оборотам") is not None and base.get("Изменение остатка") is not None:
                base["Расхождение сверки"] = base["Выручка по оборотам"] - base["Изменение остатка"]
    except Exception:
        pass

    try:
        if base.get("Остаток входящий") is not None and base.get("Остаток исходящий") is not None and base["Остаток исходящий"] < base["Остаток входящий"]:
            logger.warning("Подозрительные остатки в %s: исходящий < входящий (%s < %s)", path.name, base["Остаток исходящий"], base["Остаток входящий"])
    except Exception:
        pass

    logger.info("Результат: %s | входящий=%s исходящий=%s обороты_дт=%s обороты_кт=%s выручка=%s",
                path.name, base.get("Остаток входящий"), base.get("Остаток исходящий"),
                base.get("Обороты Дт"), base.get("Обороты Кт"), base.get("Выручка по оборотам"))
    return base

def collect_and_save(src: Path, dst: Path) -> Path:
    results=[]
    tx_sheets={}
    for p in src.iterdir():
        if not p.is_file(): continue
        plow = p.name.lower()
        if plow.startswith("сводка") or plow.startswith("summary") or (plow.endswith(".csv") and "summary_" in plow):
            logger.info("Пропускаю итоговый файл: %s", p.name); continue
        if p.suffix.lower() not in (".xlsx",".xlsm",".xltx",".xltm",".xls"): continue
        try:
            if p.stat().st_size < 1024:
                logger.info("Пропускаю малый файл (плейсхолдер): %s", p.name); continue
        except Exception: pass
        r = extract_from_file(p)
        if r.get("Транзакции найдены") and isinstance(r.get("Транзакции_DF"), pd.DataFrame):
            fname = Path(r["Файл"]).stem
            safe = re.sub(r'[\\/*?:\[\]]+', "_", fname)[:28]
            idx=1; sheet_name=safe
            while sheet_name in tx_sheets:
                idx+=1; sheet_name=f"{safe}_{idx}"
            tx_sheets[sheet_name]= r["Транзакции_DF"]
            r["Лист транзакций"]=sheet_name
            r.pop("Транзакции_DF", None)
        results.append(r)

    df = pd.DataFrame(results)
    # привести числовые колонки к числам
    num_cols = ["Остаток входящий","Остаток исходящий","Обороты Дт","Обороты Кт","Выручка по оборотам","Изменение остатка","Расхождение сверки"]
    for col in num_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    # сформировать отдельные таблицы для входящих и исходящих остатков
    df_in = df.loc[df["Остаток входящий"].notna(), ["Файл","Лист","Номер счета","Период","Остаток входящий","Выручка по оборотам","Изменение остатка","Расхождение сверки"]].copy()
    df_in.rename(columns={"Остаток входящий":"Остаток входящий"}, inplace=True)

    df_out = df.loc[df["Остаток исходящий"].notna(), ["Файл","Лист","Номер счета","Период","Остаток исходящий","Выручка по оборотам","Изменение остатка","Расхождение сверки"]].copy()
    df_out.rename(columns={"Остаток исходящий":"Остаток исходящий"}, inplace=True)

    dst.mkdir(parents=True, exist_ok=True)
    ts = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
    xfile = dst / f"Сводка_{ts}.xlsx"
    csvfile = dst / f"Сводка_{ts}.csv"

    with pd.ExcelWriter(xfile, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Сводка", index=False)
        # отдельные листы по остаткам
        df_in.to_excel(writer, sheet_name="Остатки_входящие", index=False)
        df_out.to_excel(writer, sheet_name="Остатки_исходящие", index=False)
        # транзакционные листы
        for sheet_name, df_tx in tx_sheets.items():
            try:
                df_tx.to_excel(writer, sheet_name=sheet_name, index=False)
            except Exception:
                df_tx.to_excel(writer, sheet_name=sheet_name[:25], index=False)

    df.to_csv(csvfile, index=False, encoding="utf-8")
    logger.info("Сохранено: %s и %s", xfile, csvfile)

    try:
        if sys.platform.startswith("win"):
            os.startfile(str(xfile))
        elif sys.platform.startswith("darwin"):
            os.system(f"open '{xfile}'")
        else:
            os.system(f"xdg-open '{xfile}'")
    except Exception as e:
        logger.warning("Не открылся результ.файл автоматически: %s", e)

    return xfile

def choose_and_run():
    root=tk.Tk(); root.withdraw(); root.attributes("-topmost", True)
    messagebox.showinfo("Выбор папки","Выберите папку с исходными Excel файлами (только верхний уровень).")
    src = filedialog.askdirectory(title="Папка с исходными Excel файлами")
    if not src: messagebox.showwarning("Отменено","Исходная папка не выбрана."); return
    messagebox.showinfo("Выбор папки","Выберите папку для сохранения сводки.")
    dst = filedialog.askdirectory(title="Папка для сохранения сводки")
    if not dst: messagebox.showwarning("Отменено","Целевая папка не выбрана."); return
    srcp = Path(src); dstp = Path(dst)
    logger.info("Source: %s | Destination: %s", srcp, dstp)
    result = collect_and_save(srcp, dstp)
    messagebox.showinfo("Готово", f"Сводка сохранена и открыта: {result}\nЛог: {Path(LOG_FILENAME).resolve()}")

if __name__ == "__main__":
    choose_and_run()
